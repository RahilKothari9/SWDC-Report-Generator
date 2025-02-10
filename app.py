from flask import Flask, render_template, request, redirect, url_for, session, send_file, jsonify
import pandas as pd
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.drawing.image import Image
from io import BytesIO

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.secret_key = '3d6f45a5fc12445dbac2f59c3b6c7cb1'
app.config['ALLOW_CLASS_SHARING'] = False

# Ensure the upload directory exists
if not os.path.exists(app.config['UPLOAD_FOLDER']):
    os.makedirs(app.config['UPLOAD_FOLDER'])

# Classrooms and capacities (unchanged)
classrooms = {
    'B101': {'capacity': 30, 'status': 'enabled'},
    'B102': {'capacity': 30, 'status': 'enabled'},
    'B103': {'capacity': 30, 'status': 'enabled'},
    'B104': {'capacity': 30, 'status': 'enabled'},
    'B105': {'capacity': 30, 'status': 'enabled'},
    'B201': {'capacity': 30, 'status': 'enabled'},
    'B202': {'capacity': 30, 'status': 'enabled'},
    'B203': {'capacity': 30, 'status': 'enabled'},
    'B204': {'capacity': 30, 'status': 'enabled'},
    'B205': {'capacity': 30, 'status': 'enabled'},
    'B301': {'capacity': 30, 'status': 'enabled'},
    # Add or edit classrooms as needed
}

# Updated subjects list with full details
subjects = [
    # Semester 4 (SUV2023)
    {"name": "Analysis of Algorithms", "course_code": "216U01C402", "semester": 4, "scheme": "SUV2023"},
    {"name": "Relational Database Management Systems", "course_code": "216U01C403", "semester": 4, "scheme": "SUV2023"},
    {"name": "Operating Systems", "course_code": "216U01C404", "semester": 4, "scheme": "SUV2023"},
    # Semester 6 (SUV2020)
    {"name": "Digital Signal Image Processing", "course_code": "116U01C601", "semester": 6, "scheme": "SUV2020"},
    {"name": "Information Security", "course_code": "116U01C602", "semester": 6, "scheme": "SUV2020"},
    {"name": "Artificial Intelligence", "course_code": "116U01C603", "semester": 6, "scheme": "SUV2020"},
    {"name": "Compiler Construction", "course_code": "116U01E621", "semester": 6, "scheme": "SUV2020"},
    {"name": "Data Mining and Business Intelligence", "course_code": "116U01E622", "semester": 6, "scheme": "SUV2020"},
    {"name": "Software Testing and Quality Assurance", "course_code": "116U01E623", "semester": 6, "scheme": "SUV2020"},
    {"name": "Wireless Sensor Networks and IOT", "course_code": "116U01E624", "semester": 6, "scheme": "SUV2020"},
    {"name": "Mobile Communication and Adhoc Networks", "course_code": "116U01E625", "semester": 6, "scheme": "SUV2020"},
    {"name": "Machine Learning", "course_code": "116U01E626", "semester": 6, "scheme": "SUV2020"},
    {"name": "Foundation of Micro Services", "course_code": "116U01E627", "semester": 6, "scheme": "SUV2020"},
    {"name": "Applied Cryptography", "course_code": "116U01E628", "semester": 6, "scheme": "SUV2020"},
    {"name": "Cloud Computing", "course_code": "116U01E629", "semester": 6, "scheme": "SUV2020"},
    # Semester 8 (SUV2020)
    {"name": "Bioinformatics", "course_code": "116U01E851", "semester": 8, "scheme": "SUV2020"},
    {"name": "Advanced Algorithms: Design and Analysis", "course_code": "116U01E852", "semester": 8, "scheme": "SUV2020"},
    {"name": "Internet of Everything", "course_code": "116U01E853", "semester": 8, "scheme": "SUV2020"},
    {"name": "Deep Learning", "course_code": "116U01E854", "semester": 8, "scheme": "SUV2020"},
    {"name": "Business Analytics", "course_code": "116U01E855", "semester": 8, "scheme": "SUV2020"},
    {"name": "Game Programming", "course_code": "116U01E861", "semester": 8, "scheme": "SUV2020"},
    {"name": "IoT Security", "course_code": "116U01E862", "semester": 8, "scheme": "SUV2020"},
    {"name": "Cyber Physical Systems", "course_code": "116U01E863", "semester": 8, "scheme": "SUV2020"},
    {"name": "Natural Language Processing", "course_code": "116U01E864", "semester": 8, "scheme": "SUV2020"},
    {"name": "High Performance Computing", "course_code": "116U01E865", "semester": 8, "scheme": "SUV2020"},
    {"name": "Blockchain Architecture and Application Development", "course_code": "116U01E866", "semester": 8, "scheme": "SUV2020"}
]

# Create an empty DataFrame for student data
students_df = pd.DataFrame(columns=[
    'Year', 'Programme', 'Semester', 'Student Roll', 'Name', 
    'Course Code 1', 'Course Code 2', 'Course Code 3', 'Course Code 4', 
    'Course Code 5', 'Course Code 6', 'Course Code 7', 'Course Code 8', 
    'Course Code 9', 'Course Code 10'
])

def allocate_students_to_classrooms(students, classrooms, selected_subjects):
    allocation = {}
    available_classrooms = [room for room, details in classrooms.items() if details['status'] == 'enabled']

    for subject in selected_subjects:
        # Lookup subject info by name
        subject_info = next((s for s in subjects if s['name'] == subject), None)
        if subject_info is None:
            continue
        course_code = subject_info['course_code']

        subject_students = students[students.apply(
            lambda x: course_code in x[['Course Code 1', 'Course Code 2', 'Course Code 3', 'Course Code 4', 'Course Code 5',
                                        'Course Code 6', 'Course Code 7', 'Course Code 8', 'Course Code 9', 'Course Code 10']].values,
            axis=1
        )]

        if subject_students.empty:
            continue

        total_students = len(subject_students)
        allocated_students = 0

        while allocated_students < total_students:
            if not available_classrooms:
                raise Exception(f"Not enough classrooms to allocate all students for {subject}")
            current_classroom = available_classrooms.pop(0)
            capacity = classrooms[current_classroom]['capacity']
            students_to_allocate = min(capacity, total_students - allocated_students)

            allocation[current_classroom] = {
                'subject': subject,  # subject name stored here
                'students': subject_students.iloc[allocated_students:allocated_students + students_to_allocate]
            }
            allocated_students += students_to_allocate

            # If sharing is enabled and there's leftover capacity, create a new classroom with the leftover.
            if app.config.get('ALLOW_CLASS_SHARING', False) and (students_to_allocate < capacity):
                new_class = current_classroom + "x"
                available_classrooms.insert(0, new_class)
                classrooms[new_class] = {'capacity': capacity - students_to_allocate, 'status': 'enabled'}

            # Continue to the next classroom if needed
            if allocated_students < total_students:
                continue
            elif available_classrooms:
                break
            else:
                return allocation

    return allocation

def create_seating_plan_excel(seating_plan_data):
    wb = Workbook()
    building_data = {}
    for entry in seating_plan_data:
        # For display purposes, strip any trailing "x"
        base_classroom = entry['classroom'].rstrip('x')
        building = base_classroom[0]
        if building not in building_data:
            building_data[building] = []
        # Save the stripped classroom name in the entry for display
        entry['classroom'] = base_classroom
        building_data[building].append(entry)

    # Create separate sheets for each building
    for building, data in building_data.items():
        ws = wb.create_sheet(title=f"{building}")
    
        font1 = Font(name='Times New Roman', size=15, bold=True)
        font2 = Font(name='Times New Roman', size=12, bold=True)
        font3 = Font(name='Times New Roman', size=14)
        font4 = Font(name='Times New Roman', size=28, bold=True)
        alignment1 = Alignment(horizontal='center', vertical='center', wrap_text=True)
        alignment2 = Alignment(horizontal='left', vertical='center')
        border_style = Border(left=Side(style='thin'), right=Side(style='thin'),
                              top=Side(style='thin'), bottom=Side(style='thin'))
        exam_date = datetime.now().strftime("%d.%m.%Y")
        exam_day = datetime.now().strftime("%A")

        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 13
        ws.column_dimensions['C'].width = 6
        ws.column_dimensions['D'].width = 5
        ws.column_dimensions['E'].width = 13
        ws.column_dimensions['F'].width = 17
        ws.column_dimensions['G'].width = 17
        ws.column_dimensions['H'].width = 11
        ws.column_dimensions['I'].width = 9
        ws.column_dimensions['J'].width = 7
        ws.column_dimensions['K'].width = 7
        
        title_text = "SOMAIYA VIDYAVIHAR UNIVERSITY"
        ws['B2'] = title_text
        ws['B2'].font = font1
        ws['B2'].alignment = alignment1
        ws.merge_cells('B2:K2')
        ws.row_dimensions[2].height = 20

        subtitle_text = "K.J. Somaiya School of Engineering"
        ws['B4'] = subtitle_text
        ws['B4'].font = font1
        ws['B4'].alignment = alignment1
        ws.merge_cells('B4:K4')
        ws.row_dimensions[4].height = 20

        subtitle_text = "Seating Arrangement"
        ws['B5'] = subtitle_text
        ws['B5'].font = font1
        ws['B5'].alignment = alignment1
        ws.merge_cells('B5:K5')
        ws.row_dimensions[5].height = 20

        subtitle_text = "November December 2024 Examination"
        ws['B6'] = subtitle_text
        ws['B6'].font = font1
        ws['B6'].alignment = alignment1
        ws.merge_cells('B6:K6')
        ws.row_dimensions[6].height = 20

        subtitle_text = f"Day/Date: {exam_day} / {exam_date}            Session: Afternoon             Time: 02.30 am to 05.30 pm"
        ws['B9'] = subtitle_text
        ws['B9'].font = font2
        ws['B9'].alignment = alignment1
        ws.merge_cells('B9:K9')
        ws.row_dimensions[9].height = 30

        headers = ["Programme", "Class", "Sem", "Course/ Subject", "Exam seat No.", "", "Total No. of Students",
                   "Block No.", "Floor", "BLDG"]
        ws['F11'] = 'From'
        ws['G11'] = 'To'
        for col_num, header in enumerate(headers, start=2):
            cell = ws.cell(row=10, column=col_num, value=header)
        for row in range(10, 12):
            for col in range(2, 12):
                cell = ws.cell(row=row, column=col)
                cell.border = border_style
                cell.font = font2
                cell.alignment = alignment1
        ws.merge_cells('B10:B11')
        ws.merge_cells('C10:C11')
        ws.merge_cells('D10:D11')
        ws.merge_cells('E10:E11')
        ws.merge_cells('F10:G10')
        ws.merge_cells('H10:H11')
        ws.merge_cells('I10:I11')
        ws.merge_cells('J10:J11')
        ws.merge_cells('K10:K11')
        ws.row_dimensions[11].height = 35

        start_row = 12
        prev_values = [None] * 4
        merge_starts = [None] * 4
        floor_merge_start = None
        building_merge_start = None
        prev_floor = None
        prev_building = None
        
        for idx, entry in enumerate(data, start=1):
            row = start_row + idx - 1
            current_values = [
                entry.get('programme', ''),
                entry.get('year', ''),
                entry.get('semester', ''),
                entry['subject'],
            ]
            
            ws.row_dimensions[row].height = 40

            for i in range(4):
                if current_values[i] != prev_values[i]:
                    if merge_starts[i] is not None and merge_starts[i] < row - 1:
                        ws.merge_cells(start_row=merge_starts[i], start_column=i+2, 
                                       end_row=row-1, end_column=i+2)
                    merge_starts[i] = row
                    cell = ws.cell(row=row, column=i+2, value=current_values[i])
                    cell.border = border_style
                    cell.font = font3
                    cell.alignment = alignment1
                prev_values[i] = current_values[i]

            # Use the stripped classroom for roll range and floor/building details.
            base_classroom = entry['classroom']
            ws.cell(row=row, column=6, value=entry['roll_range'].split(' - ')[0])
            ws.cell(row=row, column=7, value=entry['roll_range'].split(' - ')[1])
            ws.cell(row=row, column=8, value=entry['num_students'])
            ws.cell(row=row, column=9, value=base_classroom[:4])
            
            # Use base_classroom for floor and building.
            current_floor = base_classroom[1]   
            if current_floor != prev_floor:
                if floor_merge_start is not None:
                    ws.merge_cells(start_row=floor_merge_start, start_column=10, 
                                   end_row=row-1, end_column=10)
                floor_merge_start = row
                cell = ws.cell(row=row, column=10, value=current_floor)
                cell.border = border_style
                cell.font = font3
                cell.alignment = alignment1
            prev_floor = current_floor

            current_building = base_classroom[0]
            if current_building != prev_building:
                if building_merge_start is not None:
                    ws.merge_cells(start_row=building_merge_start, start_column=11, 
                                   end_row=row-1, end_column=11)
                building_merge_start = row
                cell = ws.cell(row=row, column=11, value=current_building)
                cell.border = border_style
                cell.font = font4
                cell.alignment = alignment1
            prev_building = current_building

            for col in range(6, 10):
                cell = ws.cell(row=row, column=col)
                cell.border = border_style
                cell.font = font3
                cell.alignment = alignment1

        for i in range(4):
            if merge_starts[i] is not None and merge_starts[i] < row:
                ws.merge_cells(start_row=merge_starts[i], start_column=i+2, 
                               end_row=row, end_column=i+2)

        if floor_merge_start is not None:
            ws.merge_cells(start_row=floor_merge_start, start_column=10, 
                           end_row=row, end_column=10)
        if building_merge_start is not None:
            ws.merge_cells(start_row=building_merge_start, start_column=11, 
                           end_row=row, end_column=11)
        
        total_height = sum(ws.row_dimensions[r].height for r in range(start_row, row + 1))
        
        static_folder = app.static_folder
        image_path = os.path.join(static_folder, 'Picture2.png')
        img = Image(image_path)
        img.height = total_height * 1.33 + 350
        img.width = 43
        ws.add_image(img, 'A1')
        image_path = os.path.join(static_folder, 'Picture1.png')
        img = Image(image_path)
        img.height = total_height * 1.33 + 350
        img.width = 27
        ws.add_image(img, 'A1')

        if total_height > 500:
            full_building_names = {
                'A': 'ARYABHAT',
                'B': 'BHASKARACHARYA'
            }
            for r in range(start_row, row + 1):
                cell = ws.cell(row=r, column=11)
                if cell.value in ['A', 'B']:
                    cell.value = full_building_names[cell.value]

        footer_text = f"Date: {exam_date}"
        ws[f'B{row+3}'] = footer_text
        ws[f'B{row+3}'].font = font2
        ws[f'B{row+3}'].alignment = alignment2
        footer_text = "EIC"
        ws[f'I{row+3}'] = footer_text
        ws[f'I{row+3}'].font = font2
        ws[f'I{row+3}'].alignment = alignment2

    wb.remove(wb['Sheet'])
    return wb

def create_roll_call_excel(classroom_data):
    wb = Workbook()
    
    for classroom, data in classroom_data.items():
        # For display, strip any trailing "x" from the classroom name.
        display_classroom = classroom.rstrip('x')
        ws = wb.create_sheet(title=f"Classroom {classroom}")

        ws.column_dimensions['A'].width = 7
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 16
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 5
        ws.column_dimensions['G'].width = 5
        ws.column_dimensions['H'].width = 5

        border_style = Border(left=Side(style='thin'), right=Side(style='thin'),
                              top=Side(style='thin'), bottom=Side(style='thin'))
        font1 = Font(name='Times New Roman', size=18, bold=True)
        font2 = Font(name='Arial', size=10, bold=True)
        font3 = Font(name='Calibri', size=11, bold=False)
        font4 = Font(name='Arial', size=10, bold=False)
        alignment1 = Alignment(horizontal='center', vertical='center')
        alignment2 = Alignment(horizontal='left', vertical='center')
        alignment3 = Alignment(horizontal='left', vertical='center', wrap_text=True)
        alignment4 = Alignment(horizontal='right', vertical='center')
        alignment5 = Alignment(horizontal='center', vertical='center', wrap_text=True)
        exam_date = datetime.now().strftime("%d.%m.%Y")

        title_text = "Somaiya Vidyavihar University"
        ws['A1'] = title_text
        ws['A1'].font = font1
        ws['A1'].alignment = alignment1
        ws.merge_cells('A1:H1')

        subtitle_text = "K.J. Somaiya School of Engineering"
        ws['A3'] = subtitle_text
        ws['A3'].font = font1
        ws['A3'].alignment = alignment1
        ws.merge_cells('A3:H3')

        text1 = "ATTENDANCE OF CANDIDATES WHO ARE PRESENT FOR THE EXAMINATION NOV-DEC 2024        School  Code: 16"
        ws['A6'] = text1
        ws['A6'].font = font2
        ws['A6'].alignment = alignment1
        ws.merge_cells('A6:H6')

        text1 = " Instructions:  Junior Supervisors should personally obtain the signature of the candidate while checking the Hall-\nTickets/ Fee Receipt / Identity Card."
        ws['A8'] = text1
        ws['A8'].font = font2
        ws['A8'].alignment = alignment3
        ws.merge_cells('A8:H8')
        ws.row_dimensions[8].height = 30

        text1 = "Supervisor’s No."
        ws['A9'] = text1
        ws['A9'].font = font2
        ws['A9'].alignment = alignment2
        ws.merge_cells('A9:B9')
        text1 = "Block No."
        ws['E9'] = text1
        ws['E9'].font = font2
        ws['E9'].alignment = alignment2
        text1 = f"{display_classroom[0:4]}"
        ws['F9'] = text1
        ws['F9'].font = font2
        ws['F9'].alignment = alignment2
        ws['F9'].border = border_style
        ws.row_dimensions[9].height = 20

        text1 = f"Programme: {data['programme']}"
        ws['A10'] = text1
        ws['A10'].font = font2
        ws['A10'].alignment = alignment2
        ws.merge_cells('A10:D10')
        text1 = f"Semester: {data['semester']}"
        ws['E10'] = text1
        ws['E10'].font = font2
        ws['E10'].alignment = alignment2
        ws.merge_cells('E10:G10')
        ws.row_dimensions[10].height = 20

        text1 = "Seat No. From: " 
        ws['A11'] = text1
        ws['A11'].font = font2
        ws['A11'].alignment = alignment2
        ws.merge_cells('A11:B11')
        text1 = f"{data['students'][0]['Student Roll']}"
        ws['C11'] = text1
        ws['C11'].font = font2
        ws['C11'].alignment = alignment2
        ws['C11'].border = border_style
        text1 = "Seat No. Upto: " 
        ws['D11'] = text1
        ws['D11'].font = font2
        ws['D11'].alignment = alignment4
        text1 = f"{data['students'][-1]['Student Roll']}"
        ws['E11'] = text1
        ws['E11'].font = font2
        ws['E11'].alignment = alignment2
        ws['E11'].border = border_style
        text1 = "Total: "
        ws['F11'] = text1
        ws['F11'].font = font2
        ws['F11'].alignment = alignment2
        ws['G11'] = len(data['students'])
        ws['G11'].font = font2
        ws['G11'].alignment = alignment2
        ws['G11'].border = border_style
        ws.row_dimensions[11].height = 20

        text1 = f"Course (Paper) Name: {data['subject_code']}"
        ws['A12'] = text1
        ws['A12'].font = font2
        ws['A12'].alignment = alignment2
        text1 = "Time: 2.30 PM - 5.30 PM"
        ws['E12'] = text1
        ws['E12'].font = font2
        ws['E12'].alignment = alignment2
        ws.row_dimensions[12].height = 20
        
        text1 = f"Date : {exam_date}"
        ws['A13'] = text1
        ws['A13'].font = font2
        ws['A13'].alignment = alignment2
        text1 = "Session:  Afternoon"
        ws['D13'] = text1
        ws['D13'].font = font2
        ws['D13'].alignment = alignment2
        text1 = "Section: "
        ws['E13'] = text1
        ws['E13'].font = font2
        ws['E13'].alignment = alignment2
        ws.row_dimensions[13].height = 20

        headers = ["SRNO","SEAT NO.", "ANSWERBOOK SR.NO.", "NAME", "CANDIDATE SIGNATURE","RECORD OF SUPPLEMENTS"]
        ws.row_dimensions[14].height = 30
        start_row = 14
        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=col_num, value=header)
            cell.font = font2
            cell.alignment = alignment5
            cell.border = border_style
        ws.merge_cells(start_row=start_row, start_column=6, end_row=start_row, end_column=8)
        record_of_supplements_cell = ws.cell(row=start_row, column=6)
        record_of_supplements_cell.value = "RECORD OF SUPPLEMENTS"
        record_of_supplements_cell.font = font2
        record_of_supplements_cell.alignment = alignment5
        record_of_supplements_cell.border = border_style

        for idx, student in enumerate(data['students'], start=start_row + 1):
            ws.append([idx - start_row, student['Student Roll'], "", student['Name'], "", ""])
            ws.row_dimensions[idx].height = 23
            for col_num in range(1, 9):
                cell = ws.cell(row=idx, column=col_num)
                cell.font = font3
                cell.alignment = alignment1
                cell.border = border_style

        static_folder = app.static_folder
        image_path = os.path.join(static_folder, 'absentees.png')
        img = Image(image_path)
        num = len(data['students']) + 16
        ws.add_image(img, f'B{num}')

        num = len(data['students']) + 26
        text1 = "Total Number of Candidates Allotted  to the block: "
        ws[f'A{num}'] = text1
        ws[f'A{num}'].font = font2
        ws[f'A{num}'].alignment = alignment2
        ws.merge_cells(f'A{num}:D{num}')

        num = len(data['students']) + 28
        text1 = "Total Number of Candidates Present: "
        ws[f'A{num}'] = text1
        ws[f'A{num}'].font = font2
        ws[f'A{num}'].alignment = alignment2
        ws.merge_cells(f'A{num}:D{num}')

        num = len(data['students']) + 30
        text1 = "Total Number of Candidates Absent: "
        ws[f'A{num}'] = text1
        ws[f'A{num}'].font = font2
        ws[f'A{num}'].alignment = alignment2
        ws.merge_cells(f'A{num}:D{num}')
        
        #footer
        ws[f"A{num+13}"] = "Checked by:"
        ws[f"D{num+13}"] = "Name and Signature of Block Supervisor: "
        ws[f"A{num+15}"] = "Date:"
        ws[f"D{num+15}"] = "Signature of Sr. Supervisor: "
        ws[f'A{num+13}'].font = font4
        ws[f'A{num+13}'].alignment = alignment2
        ws[f'D{num+13}'].font = font4
        ws[f'D{num+13}'].alignment = alignment2
        ws[f'A{num+15}'].font = font4
        ws[f'A{num+15}'].alignment = alignment2
        ws[f'D{num+15}'].font = font4
        ws[f'D{num+15}'].alignment = alignment2
       
    wb.remove(wb['Sheet'])
    return wb

@app.route('/', methods=['GET'])
def index():
    return render_template('index.html', subjects=subjects)

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return redirect(url_for('index'))
    file = request.files['file']
    if file.filename == '':
        return redirect(url_for('index'))
    if file and file.filename.endswith('.xlsx'):
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
        file.save(filepath)
        global students_df
        students_df = pd.read_excel(filepath)
        session['file_uploaded'] = True
        return redirect(url_for('index'))
    return redirect(url_for('index'))

@app.route('/select_subjects', methods=['POST'])
def select_subjects():
    selected_subjects = request.form.getlist('subjects')
    session['selected_subjects'] = selected_subjects
    return redirect(url_for('index'))

@app.route('/seating_plan')
def seating_plan():
    if 'file_uploaded' not in session or 'selected_subjects' not in session:
        return redirect(url_for('index'))
    selected_subjects = session['selected_subjects']

    try:
        allocation = allocate_students_to_classrooms(students_df, classrooms, selected_subjects)
        seating_plan_data = []
        for classroom, data in allocation.items():
            students = data['students']
            roll_numbers = students['Student Roll'].tolist()
            # Look up the subject info from the subjects list using subject name.
            subject_info = next((s for s in subjects if s['name'] == data['subject']), {})
            seating_plan_data.append({
                'classroom': classroom.rstrip('x'),  # Strip trailing x's for display
                'subject': subject_info.get('name', data['subject']),
                'semester': subject_info.get('semester', students['Semester'].iloc[0] if not students.empty else ''),
                'scheme': subject_info.get('scheme', ''),
                'num_students': len(students),
                'roll_range': f"{min(roll_numbers)} - {max(roll_numbers)}"
            })
        return render_template('seating_plan.html', seating_plan=seating_plan_data)
    except Exception as e:
            error_message = str(e)  
            return render_template('error.html', message=error_message), 400
    
@app.route('/download_seating_plan')
def download_seating_plan():
    if 'file_uploaded' not in session or 'selected_subjects' not in session:
        return redirect(url_for('index'))

    selected_subjects = session['selected_subjects']
    allocation = allocate_students_to_classrooms(students_df, classrooms, selected_subjects)

    seating_plan_data = []
    for classroom, data in allocation.items():
        students = data['students']
        roll_numbers = students['Student Roll'].tolist()
        seating_plan_data.append({
            'classroom': classroom.rstrip('x'),
            'subject': data['subject'],
            'num_students': len(students),
            'programme': students['Programme'].iloc[0] if not students.empty else '',
            'year': students['Year'].iloc[0] if not students.empty else '',
            'semester': students['Semester'].iloc[0] if not students.empty else '',
            'roll_range': f"{min(roll_numbers)} - {max(roll_numbers)}"
        })

    wb = create_seating_plan_excel(seating_plan_data)
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, download_name='seating_plan.xlsx', as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/roll_call_list')
def roll_call_list():
    if 'file_uploaded' not in session or 'selected_subjects' not in session:
        return redirect(url_for('index'))
    selected_subjects = session['selected_subjects']

    try:
        allocation = allocate_students_to_classrooms(students_df, classrooms, selected_subjects)
        classroom_data = {}
        for classroom, data in allocation.items():
            subject_info = next((subject for subject in subjects if subject['name'] == data['subject']), None)
            if subject_info is None:
                continue
            # Save display_classroom as the classroom name stripped of trailing x's.
            display_classroom = classroom.rstrip('x')
            classroom_data[display_classroom] = {
                'subject_name': subject_info['name'],
                'subject_code': subject_info['course_code'],
                'scheme': subject_info['scheme'],
                'programme': data['students'].iloc[0]['Programme'] if not data['students'].empty else '',
                'semester': subject_info['semester'],
                'students': data['students'].to_dict(orient='records')
            }
        exam_date = datetime.now().strftime("%d.%m.%Y")
        exam_time = "2.30 PM - 5.30 PM"
        exam_session = "Afternoon"
        return render_template('roll_call_list.html',
                            classroom_data=classroom_data,
                            exam_date=exam_date,
                            exam_time=exam_time,
                            exam_session=exam_session)
    except Exception as e:
                error_message = str(e)  
                return render_template('error.html', message=error_message), 400

@app.route('/download_roll_call_list')
def download_roll_call_list():
    if 'file_uploaded' not in session or 'selected_subjects' not in session:
        return redirect(url_for('index'))
    
    selected_subjects = session['selected_subjects']
    allocation = allocate_students_to_classrooms(students_df, classrooms, selected_subjects)

    classroom_data = {}
    for classroom, data in allocation.items():
        subject_info = next((subject for subject in subjects if subject['name'] == data['subject']), None)
        if subject_info is None:
            continue
        display_classroom = classroom
        classroom_data[display_classroom] = {
            'subject_name': subject_info['name'],
            'subject_code': subject_info['course_code'],
            'scheme': subject_info['scheme'],
            'programme': data['students'].iloc[0]['Programme'] if not data['students'].empty else '',
            'semester': subject_info['semester'],
            'students': data['students'].to_dict(orient='records')
        }

    wb = create_roll_call_excel(classroom_data)
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, download_name='roll_call_list.xlsx', as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/configure', methods=['GET', 'POST'])
def configure():
    global classrooms, subjects

    if request.method == 'POST':
        # Check if the form submission is for updating the share option.
        if 'allow_class_sharing' in request.form:
            # The checkbox sends 'on' when checked; if not present, assume False.
            app.config['ALLOW_CLASS_SHARING'] = request.form.get('allow_class_sharing') == 'on'
        
        # Handle adding classrooms
        if 'classroom' in request.form and 'capacity' in request.form:
            classroom = request.form['classroom']
            capacity = int(request.form['capacity'])
            classrooms[classroom] = {'capacity': capacity, 'status': 'enabled'}
        
        # Handle adding subjects
        elif 'name' in request.form and 'course_code' in request.form and 'semester' in request.form:
            name = request.form['name']
            course_code = request.form['course_code']
            semester = int(request.form['semester'])
            scheme = request.form.get('scheme', '')
            subjects.append({"name": name, "course_code": course_code, "semester": semester, "scheme": scheme})
            
    # Pass the current option value to the template.
    return render_template('configure.html', classrooms=classrooms, subjects=subjects, allow_class_sharing=app.config['ALLOW_CLASS_SHARING'])

@app.route('/delete_classroom', methods=['POST'])
def delete_classroom():
    classroom_to_delete = request.json.get('classroom')
    if classroom_to_delete in classrooms:
        del classrooms[classroom_to_delete]
        return jsonify({'status': 'success', 'message': f'Classroom {classroom_to_delete} deleted successfully'})
    return jsonify({'status': 'error', 'message': f'Classroom {classroom_to_delete} not found'})

@app.route('/delete_subject', methods=['POST'])
def delete_subject():
    course_code_to_delete = request.json.get('course_code')
    global subjects
    subjects = [s for s in subjects if s['course_code'] != course_code_to_delete]
    return jsonify({'status': 'success', 'message': f'Subject {course_code_to_delete} deleted successfully'})

@app.route('/toggle_classroom_status', methods=['POST'])
def toggle_classroom_status():
    classroom = request.json.get('classroom')
    status = request.json.get('status')
    if classroom in classrooms:
        classrooms[classroom]['status'] = 'enabled' if status == 'enabled' else 'disabled'
        return jsonify({'status': 'success', 'message': f'Classroom {classroom} {status}d successfully'})
    return jsonify({'status': 'error', 'message': f'Classroom {classroom} not found'})

@app.route('/modify_classroom_capacity', methods=['POST'])
def modify_classroom_capacity():
    classroom = request.json.get('classroom')
    capacity = request.json.get('capacity')
    if classroom in classrooms:
        classrooms[classroom]['capacity'] = int(capacity)
        return jsonify({'status': 'success', 'message': f'Classroom {classroom} capacity modified successfully'})
    return jsonify({'status': 'error', 'message': f'Classroom {classroom} not found'})

if __name__ == '__main__':
    app.run(debug=True)
